# -*- coding: utf-8 -*-
"""
Created on Fri Mar 24 16:09:01 2023


"""

from pywinauto.application import Application
import imaplib
import email
from openpyxl import Workbook
import time
import pyautogui as auto

# Load the user name and passwd 
user, password = "udaysankar2003@gmail.com","yxoqclmsdmqsjxeg"

# URL for IMAP connection
imap_url = 'imap.gmail.com'

# Connection with GMAIL using SSL
my_mail = imaplib.IMAP4_SSL(imap_url)

# Log in using your credentials
my_mail.login(user, password)

# Select the Inbox to fetch unread messages
my_mail.select('Inbox')

# Define Key and Value for email search
key = "SUBJECT"
value = "New_Registration"
_, data = my_mail.search(None, 'SEEN', key, value)  # Search for unread emails with specific key and value

mail_id_list = data[0].split()  # IDs of all unread emails that we want to fetch 

msgs = [] # empty list to capture all messages
# Iterate through messages and extract data into the msgs list
for num in mail_id_list:
    typ, data = my_mail.fetch(num, '(RFC822)') # RFC822 returns whole message (BODY fetches just body)
    msgs.append(data)

# Create a new Excel workbook
wb = Workbook()
ws = wb.active

# Set the column headers
ws['A1'] = 'From'
ws['B1'] = 'Subject'
ws['C1'] = 'Body'
ws['D1'] = 'PatientID'
ws['E1'] = 'PatientName'
ws['F1'] = 'PatientEmail'
ws['G1'] = 'Ph_No'
ws['H1'] = 'Blood_Grp'

# Iterate through the messages and extract the subject and body
for i, msg in enumerate(msgs[::-1]):
    for response_part in msg:
        if type(response_part) is tuple:
            my_msg = email.message_from_bytes((response_part[1]))
            From = my_msg['from']
            subject = my_msg['subject']
            body = ''
            for part in my_msg.walk():
                if part.get_content_type() == 'text/plain':
                    body = part.get_payload()
            datastr=body.split(",")
            name=datastr[0]
            age=datastr[1]
            ph_no=datastr[2]
            address=datastr[3]
            blood_grp=datastr[4]
            # Add the subject and body to the worksheet
            ws['A' + str(i+2)] = From
            ws['B' + str(i+2)] = subject
            ws['C' + str(i+2)] = body
            ws['D' + str(i+2)] = name
            ws['E' + str(i+2)] = age
            ws['F' + str(i+2)] = ph_no
            ws['G' + str(i+2)] = address
            ws['H' + str(i+2)] = blood_grp
            
            app = Application(backend='uia').start('C://Users//UDAY SANKAR//AppData\Local//Programs//Data_Entry//DATA_ENTRY.exe')
            app = Application(backend='uia').connect(title='DATA_ENTRY',timeout=20)
            #time.sleep(5)
            #app.DataEntry.print_control_identifiers()
            Maximize = app.DataEntry.child_window(title="Maximize", control_type="Button").wrapper_object()
            Maximize.click_input()

            Maximize = app.DataEntry.child_window(title="System", control_type="MenuItem").wrapper_object()
            Maximize.click_input()

            x2,y2=auto.locateCenterOnScreen(r"C:\Users\UDAY SANKAR\Desktop\codes\patientID.png",confidence=0.9)
            auto.moveTo(x2,y2,1)
            auto.click()
            time.sleep(1)
            auto.write(name)
            x3,y3=auto.locateCenterOnScreen(r"C:\Users\UDAY SANKAR\Desktop\codes\patientName.png",confidence=0.9)
            auto.moveTo(x3,y3,1)
            auto.click()
            time.sleep(1)
            auto.write(age)
            x4,y4=auto.locateCenterOnScreen(r"C:\Users\UDAY SANKAR\Desktop\codes\patientEmail.png",confidence=0.9)
            auto.moveTo(x4,y4,1)
            auto.click()
            time.sleep(1)
            auto.write(ph_no)
            x5,y5=auto.locateCenterOnScreen(r"C:\Users\UDAY SANKAR\Desktop\codes\Ph_No.png",confidence=0.9)
            auto.moveTo(x5,y5,1)
            auto.click()
            time.sleep(1)
            auto.write(address)
            x6,y6=auto.locateCenterOnScreen(r"C:\Users\UDAY SANKAR\Desktop\codes\BloodGrp.png",confidence=0.9)
            auto.moveTo(x6,y6,1)
            auto.click()
            time.sleep(1)
            auto.write(blood_grp)
            x7,y7=auto.locateCenterOnScreen(r"C:\Users\UDAY SANKAR\Desktop\codes\add.png",confidence=0.9)
            auto.moveTo(x7,y7,1)
            auto.click()

            Maximize = app.DataEntry.child_window(title="Close", control_type="Button").wrapper_object()
            Maximize.click_input()

# Mark the fetched emails as read
#for num in mail_id_list:
#    my_mail.store(num, '+FLAGS', '\\Seen')

# Save the workbook to an Excel file
wb.save('emails.xlsx')
